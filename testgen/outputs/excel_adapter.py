"""
Excel / CSV 输出适配器
------------------------
将测试套件导出为表格格式，方便非开发人员查看和评审。

ExcelAdapter 特性:
  - 每个 suite 一个 sheet（sheet 名截断至 31 字符）
  - 优先级着色: high=红色 / medium=黄色 / low=绿色
  - 同一用例多步骤时，自动合并非步骤列单元格
  - 表头: 蓝底白字加粗，自适应列宽

CSVAdapter 特性:
  - UTF-8 BOM 编码（Excel 可直接打开，中文不乱码）
  - 逗号分隔符
"""

import csv
from datetime import datetime
from pathlib import Path
from typing import Any

from ..core.base import BaseOutputAdapter
from ..core.models import GenerationContext, TestSuite


class ExcelAdapter(BaseOutputAdapter):
    """生成 Excel 格式的测试用例表格（见模块文档了解详情）"""

    def format_name(self) -> str:
        return "excel"

    def write(
        self, suites: list[TestSuite], context: GenerationContext
    ) -> list[str]:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")

        output_dir = Path(context.output_dir) / "excel"
        output_dir.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        # 保留默认 sheet，后续重命名或删除

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 优先级颜色
        priority_fills = {
            "P0": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "P2": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "P3": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        }

        headers = [
            "用例ID", "测试类型", "测试模块", "子功能", "用例标题",
            "前置条件/测试数据", "操作步骤", "所属系统", "预期结果",
        ]

        generated_files: list[str] = []

        for i, suite in enumerate(suites):
            sheet_name = suite.name[:31]
            if i == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            row = 2
            for idx, case in enumerate(suite.test_cases):
                test_id = case.id if case.id else f"id_{idx+1:03d}"
                test_type_str = case.test_type.value
                module_str = case.module or suite.name.split("_")[0] if "_" in suite.name else ""
                sub_feature_str = case.sub_feature or case.name.split(" - ")[1] if " - " in case.name else ""
                title_str = case.description or case.name
                precond_str = case.preconditions
                system_str = case.system_name or ""

                # 操作步骤
                steps_lines = []
                for step in case.steps:
                    steps_lines.append(f"{step.step_number}.{step.action}")
                steps_str = "\n".join(steps_lines)

                expected_str = case.expected_result
                if not expected_str and case.steps:
                    expected_str = case.steps[-1].expected_result

                row_data = [
                    test_id, test_type_str, module_str, sub_feature_str,
                    title_str, precond_str, steps_str, system_str, expected_str,
                ]
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                row += 1

            # 列宽调整
            col_widths = [20, 12, 15, 15, 55, 30, 45, 15, 35]
            for col, width in enumerate(col_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        # 保存
        filepath = output_dir / "test_cases.xlsx"
        wb.save(str(filepath))
        generated_files.append(str(filepath))

        return generated_files


class CSVAdapter(BaseOutputAdapter):
    """生成 CSV 格式的测试用例表格（见模块文档了解详情）"""

    def format_name(self) -> str:
        return "csv"

    def write(
        self, suites: list[TestSuite], context: GenerationContext
    ) -> list[str]:
        output_dir = Path(context.output_dir) / "csv"
        output_dir.mkdir(parents=True, exist_ok=True)

        headers = [
            "用例ID", "用例名称", "描述", "测试类型", "优先级",
            "标签", "前置条件", "整体预期结果", "步骤编号", "操作步骤", "预期结果", "断言", "期望状态码", "用例预期结果",
        ]
        generated_files: list[str] = []

        for suite in suites:
            filename = self._sanitize(suite.name) + ".csv"
            filepath = output_dir / filename

            with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(headers)

                for case in suite.test_cases:
                    preconditions_str = case.preconditions
                    tags_str = ", ".join(case.tags)

                    if case.steps:
                        for step in case.steps:
                            assertions_str = "; ".join(step.assertions)
                            writer.writerow([
                                case.id, case.name, case.description,
                                case.test_type.value, case.priority, tags_str,
                                preconditions_str, case.expected_result, step.step_number, step.action,
                                step.expected_result, assertions_str, case.expected_status,
                            ])
                    else:
                        writer.writerow([
                            case.id, case.name, case.description,
                            case.test_type.value, case.priority, tags_str,
                            preconditions_str, case.expected_result, "", "", "", "", case.expected_status,
                        ])

            generated_files.append(str(filepath))

        return generated_files

    def _sanitize(self, name: str) -> str:
        import re
        return re.sub(r'[<>:"/\\|?*]', "_", name).replace(" ", "_")
